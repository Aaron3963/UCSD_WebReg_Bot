from ast import Try, While
from time import sleep, time
from xmlrpc.client import Boolean
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select,WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

import openpyxl
bk=openpyxl.load_workbook('config.xlsx')
sheet=bk.active

maxRow=sheet.max_row
maxCol=sheet.max_column

USERNAME = sheet.cell(1,2).value
PASSWORD = sheet.cell(2,2).value
QUARTER = sheet.cell(3,2).value

dict = {}
courseChecklist = {}
for i in range(3,maxRow+1):
    list = []
    for j in range(2,maxCol+1):
        cell = str(sheet.cell(i,j).value)
        if cell == 'None':
            continue
        if len(cell) < 6:
            for numZero in range (0,6-len(cell)):
                cell = '0'+cell
        list.append(cell)
    dict[sheet.cell(i,1).value] = list
    courseChecklist [sheet.cell(i,1).value] = False
print(dict)


wd = webdriver.Chrome(service= Service(r'./chromedriver.exe'))
wd.implicitly_wait(10)

wd.get('https://act.ucsd.edu/webreg2')

username = wd.find_element(By.ID,'ssousername')
username.send_keys(USERNAME)
password = wd.find_element(By.ID,'ssopassword')
password.send_keys(PASSWORD)

wd.find_element(By.XPATH,'//*[@id="login"]/button').click()

sleep(15)

#Go page
def goPage(wd,quarter):
    wd.get('https://act.ucsd.edu/webreg2')
    termSelect = Select(wd.find_element(By.ID,'startpage-select-term'))
    
    # TODO
    termSelect.select_by_visible_text(quarter)

    wd.find_element(By.ID,'startpage-button-go').click()


def search(wd,course):
    searchBar = wd.find_element(By.ID, 's2id_autogen1')
    searchBar.clear()
    searchBar.send_keys(course)
    
    #Confirm Selection
    wd.find_element(By.ID, 'select2-drop-mask').click()

    #Search Class
    wd.find_element(By.ID, 'search-div-t-b1').click()

    #Expand Section List
    wd.find_element(By.XPATH, '//*[@id="search-div-b-tableghead_0_0"]/td/span').click()
    

def check(wd,course,id):
    sectionEnrollID = 'search-enroll-id-'+str(id)
    sectionWaitID = 'search-wait-id-'+str(id)

    wd.implicitly_wait(0.05)
    try:
        targetClass = wd.find_element(By.ID,sectionEnrollID)
    except Exception:
        targetClass = wd.find_element(By.ID,sectionWaitID)

    wd.implicitly_wait(10)

    if course not in targetClass.get_attribute('class'):
        print("Class Mismatch. Abort.")

    if 'Waitlist' == targetClass.get_attribute('value'):
        print(course+id+": No Space, retrying...")
        return False

    if 'Enroll' == targetClass.get_attribute('value'):
        #wd.find_element(By.ID)
        return True

def enroll(wd,course,id,checklist):
    sectionEnrollID = 'search-enroll-id-'+str(id)

    wd.find_element(By.ID,sectionEnrollID).click()

    targetClass = wd.find_element(By.ID,'dialog-enroll').find_element(By.CLASS_NAME, 'dialog-tip-class')

    enroll = False
   
    sleep(1) #Must use sleep for web to load new element
    error = targetClass.find_elements(By.CLASS_NAME,'error')

    # print(len(error))

    if len(error) == 0:
        if "Confirm" in targetClass.get_attribute('textContent'):
            print("No Error Detected, Proceed Enrollment")
            enroll = True

    if enroll:
        try:
            wd.find_element(By.XPATH,'/html/body/div[28]/div[3]/div/button[2]').click()
        except Exception:
            wd.find_element(By.XPATH,'/html/body/div[28]/div[3]/div/button[2]').click()

        try:
            wd.find_element(By.ID,'dialog-after-action-email').click()
        except Exception:
            wd.find_element(By.ID,'dialog-after-action-email').click()
            checklist[course] = True
            print("Successfully enrolled in "+course+id+"!")
            
    elif not enroll:
        try:
            wd.find_element(By.ID,'dialog-after-action-close').click()
        except Exception:
            wd.find_element(By.ID,'dialog-after-action-close').click()


    return enroll

    
def checkCourseList(checklist):
    for enrolled in checklist.values():
        if not enrolled:
            return False
    return True


def main():
    startTime = time()
    passedTime = startTime
    while passedTime - startTime < 30:
        for course in dict.keys():
            if courseChecklist[course]:
                continue
            search(wd,course)
            for sectionID in dict[course]:
                vacant = check(wd,course,sectionID)
                if vacant:
                    enrolled = enroll(wd,course,sectionID,courseChecklist)
                    if enrolled:
                        return True
                passedTime = time()
    return False

goPage(wd,QUARTER)
main()

while not checkCourseList(courseChecklist):
    print(courseChecklist)
    print("Refresh Page")
    goPage(wd,QUARTER)
    main()

wd.quit()